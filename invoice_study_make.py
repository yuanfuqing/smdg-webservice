from shutil import copyfile
import os
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from tkinter import filedialog
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import requests
import time
from bs4 import BeautifulSoup
from datetime import date
import requests
import random
import json
from hashlib import md5
import numpy as np
from PIL import Image
import xlsxwriter
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

align = Alignment(horizontal='left', vertical='center')
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)
date_now = time.strftime("%d/%m/%Y", time.localtime())

def decision(a):
    if (len(str(a)) == 0):
        return ''
    elif (a>=0):
        return '无'
    elif a<0:
        return '有'
def get_invoicedate(source):
    writer_1 = pd.ExcelFile(source)
    c = writer_1.sheet_names
    datainvoice = writer_1.parse(c[0])
    datainvoice = datainvoice.dropna(subset=["货箱编号"])
    # 已有产品申报单价
    datainvoice['产品申报单价'] = datainvoice['产品申报单价'].apply(lambda x: float(x))
    datainvoice['产品申报数量'] = datainvoice['产品申报数量'].apply(lambda x: int(x))
    datainvoice['货箱重量(KG)'] = datainvoice['货箱重量(KG)'].apply(lambda x: float(x))
    datainvoice['跟踪号'] = datainvoice['跟踪号'].apply(lambda x: str(x).split(".")[0])
    datainvoice['产品海关编码'] = datainvoice['产品海关编码'].apply(lambda x: str(x)[:10])
    datainvoice['产品海关编码'] = datainvoice['产品海关编码'].apply(lambda x: int(x))
    datainvoice['申报总价'] = datainvoice['产品申报单价'] * datainvoice['产品申报数量']
    
    datainvoice['毛重比例'] = datainvoice['货箱重量(KG)'] / datainvoice['货箱重量(KG)'].sum()
    datainvoice['包裹净重'] = datainvoice['货箱重量(KG)'] - len(set(datainvoice['货箱编号'].tolist()))*1*datainvoice['毛重比例']
    datainvoice['产品净重'] = ((datainvoice['包裹净重'] / datainvoice['产品申报数量'])-0.005).round(2)
    datainvoice['包裹净重'] = round(datainvoice['产品净重'] * datainvoice['产品申报数量'], 2)

    
    datainvoice['箱数'] = datainvoice['货箱编号']  # 先等于运单号，然后在调整
    datainvoice['每公斤价值'] = round(datainvoice['申报总价'] / datainvoice['货箱重量(KG)'],2)   # 先等于运单号，然后在调整
    datainvoice['产品英文品名'] = datainvoice['产品英文品名']
    datainvoice['产品中文品名'] = datainvoice['产品中文品名']
    datainvoice = datainvoice.sort_values("货箱编号")
    datainvoice = datainvoice.fillna("")
    return datainvoice
def extrait_hscode(hscode, today):
    dic = []
    url = "https://eservices.minfin.fgov.be/extTariffBrowser/Measure?cnCode=%s&country=29422&trade=0&cssfile=tarbro" \
          "&date=%s&lang=EN&page=1" % (
              hscode, today)
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/76.0.3809.132 Safari/537.36'}
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'html.parser')
    hscode = soup.find('span', class_="smaller-title").text.replace(" ","")
    description_hscode = soup.find('ul', class_="nostyle").getText().replace('\n', '').replace(
        '                                   ', ' /')
    footnote = soup.find('table', class_="table-nopadding").getText().replace('\n', '').strip().replace("Footnotes:",
                                                                                                        "")
    supplementary_unit = soup.find('table', class_="table-nopadding bottom-aligned").getText().replace('\n',
                                                                                                       '').strip().replace(
        "Supplementary unit:", "")
    tables = soup.find_all('div', class_="meas-header")
    for table in tables:
        table_infos = table.getText().split("\n")
        type_table = table_infos[0]  # 表格类型
        table_infos_sorts = table_infos[19:]
        nb_ligne = len(table_infos_sorts) / 20
        for x in range(int(nb_ligne)):
            Geographical_area = table_infos_sorts[20 * x + 0] + "  " + table_infos_sorts[20 * x + 1]
            Measure_type = table_infos_sorts[20 * x + 2] + "  " + table_infos_sorts[20 * x + 3]
            Tariff = table_infos_sorts[20 * x + 4] + "  " + table_infos_sorts[20 * x + 5]
            dic_0 = {"type_table":type_table,
                     "Measure_type":Measure_type,
                     "Tariff":Tariff,
            "Geographical_area": Geographical_area,}
            dic.append(dic_0)
    pd_hscode_no_info = pd.DataFrame(list(dic))
    if "CN - China  " in pd_hscode_no_info["Geographical_area"].tolist():
        anti_dumping = "anti-dumping"
    else:
        anti_dumping = ""
    duty = pd_hscode_no_info["Tariff"].loc[(pd_hscode_no_info["type_table"]=="Tariff measures") & (pd_hscode_no_info["Measure_type"]=="Third country duty          ") ].tolist()[0]
    return description_hscode,anti_dumping,duty


def translate_eng_cn(query):
    # Set your own appid/appkey.
    appid = '20220629001259722'
    appkey = 'vkooiwx4xLqOl9C8NjvW'
    # For list of language codes, please refer to `https://api.fanyi.baidu.com/doc/21`
    from_lang = 'en'
    to_lang = 'zh'
    endpoint = 'http://api.fanyi.baidu.com'
    path = '/api/trans/vip/translate'
    url = endpoint + path
    # Generate salt and sign
    def make_md5(s, encoding='utf-8'):
        return md5(s.encode(encoding)).hexdigest()
    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()
    description_en_chinois = result['trans_result'][0]['dst']
    return description_en_chinois

def declaration_product(product):
    payload = {"includeUK": "false",
               "lang": "CN",
               "partner": "CN",
               "product": product,
               "years": '2021'}
    headers = {'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/102.0.0.0 Safari/537.36',
               'Cookie': ''}
    r = requests.get(
        'https://webgate.ec.europa.eu/flows/public/v1/stats?', params=payload, headers=headers)
    list_value = r.json()['rows']
    importValue_total, importQuantity_total = 0, 0
    for value in list_value:
        country = value['country']
        importValue = value['samples']['2021']['importValue']
        if len(str(importValue).split(".")[-1]) == 3:
            importValue = int(str(importValue).replace(".", ""))
        elif len(str(importValue).split(".")[-1]) == 2:
            importValue = int(str(importValue).replace(".", "")) * 10
        elif len(str(importValue).split(".")[-1]) == 1:
            importValue = int(str(importValue).replace(".", "")) * 100

        importQuantity = value['samples']['2021']['importQuantity']
        if len(str(importQuantity).split(".")[-1]) == 3:
            importQuantity = int(str(importQuantity).replace(".", ""))
        elif len(str(importQuantity).split(".")[-1]) == 2:
            importQuantity = int(str(importQuantity).replace(".", "")) * 10
        elif len(str(importQuantity).split(".")[-1]) == 1:
            importQuantity = int(str(importQuantity).replace(".", "")) * 100

        importValue_total, importQuantity_total = \
            importValue_total + importValue, importQuantity_total + importQuantity

    country = "EURO 27"
    if importQuantity_total == 0:
        import_kg_total = 0
    else:
        import_kg_total = round(importValue_total / importQuantity_total, 2)
    return  import_kg_total

def study_invoice(data_hscode,source,study_name):
    today = date.today()
    years = 2021
    datainvoice = get_invoicedate(source)
    df_hscode_invoice = datainvoice[["运单号","产品海关编码","每公斤价值","产品英文品名","产品中文品名"]].drop_duplicates().sort_values("产品海关编码")
    df_hscode_analyse = pd.merge(df_hscode_invoice,data_hscode,left_on="产品海关编码",right_on="hscode",how='left')
    list_hscode_no_info = set(df_hscode_analyse['产品海关编码'].loc[df_hscode_analyse['hscode'].isna()].tolist())
    if len(list_hscode_no_info) == 0:
        pass
    else:
        list_o = []
        hscode_no_exsite = []
        n = 0
        print("共计%s个海关码不再数据库，需进行海关网站抓取" % (len(list_hscode_no_info)))
        print("-------------------------")
        for hscode_on_info in list_hscode_no_info:
            n = n + 1
            print("正在提取%s个海关码 :"%(n),hscode_on_info)
            try:
                description_hscode,anti_dumping,duty = extrait_hscode(hscode_on_info, today)
                description_en_chinois = translate_eng_cn(description_hscode)
                product = str(hscode_on_info)[:8]
                import_kg_total = declaration_product(product)
                a = {'hscode':hscode_on_info,'Duty':duty,'import_euro_kg': import_kg_total,
                     'anti_dumping': anti_dumping,'description_hscode': description_hscode,
                     'description_en_chinois': description_en_chinois,
                     'date_search': today,'lien':''}
                list_o.append(a)
            except:
                b = {'hscode':hscode_on_info,'Statue':"未找到，人工核实"}
                hscode_no_exsite.append(b)
                print("****************************未找到海关码  %s   ，请核实"%(hscode_on_info))
        df_no_existe = pd.DataFrame(list(hscode_no_exsite))
        df_hscode_insert = pd.DataFrame(list(list_o))
        data_hscode = data_hscode.append(df_hscode_insert,ignore_index=True)
    df_hscode_analyse = pd.merge(df_hscode_invoice,data_hscode,left_on="产品海关编码",right_on="hscode",how='left')
    df_hscode_analyse["差值"] =  df_hscode_analyse["每公斤价值"] - df_hscode_analyse["import_euro_kg"]
    df_hscode_analyse["低报风险"] = df_hscode_analyse['差值'].apply(decision)
    df_antidumping = df_hscode_analyse[df_hscode_analyse["anti-dumping"]=="anti-dumping"]
    df_low_value = df_hscode_analyse[df_hscode_analyse["低报风险"]=="有"]
    table_df_low_value = pd.pivot_table(df_low_value,values=['import_euro_kg', '每公斤价值','差值'],index=['产品海关编码','低报风险','产品英文品名','产品中文品名'],
                           aggfunc={'import_euro_kg': np.mean,
                                    '每公斤价值': np.mean,
                                    '差值': np.mean})
    table = pd.pivot_table(df_hscode_analyse,values=['import_euro_kg', '每公斤价值'],index=['产品海关编码','description_en_chinois','产品中文品名'],
                           aggfunc={'import_euro_kg': np.mean,
                                    '每公斤价值': np.mean})
    with pd.ExcelWriter(study_name,engine="openpyxl") as writer:
            df_hscode_analyse.to_excel(writer, sheet_name='申报信息总结', index=False)
            table.to_excel(writer, sheet_name='透视表格')
            try:
                df_no_existe.to_excel(writer, sheet_name='海关码不存在', index=False)
            except:
                pass
            df_antidumping.to_excel(writer, sheet_name='反倾销', index=False)
            try:
                table_df_low_value.to_excel(writer, sheet_name='低报风险')
            except:
                pass
            try:
                df_hscode_insert.to_excel(writer, sheet_name='打包发给米西', index=False)
            except:
                pass

def get_company_information(company_info):
    writer_1 = pd.ExcelFile(company_info)
    c = writer_1.sheet_names
    for x in c:
        if "发件人" in str(x):
            datasender = writer_1.parse(x)
        elif "收件人" in str(x):
            datarecipeter = writer_1.parse(x)
    return datasender

def Merge_cells(ws, target_list, start_row, col):
    '''
    ws: 是需要操作的工作表
    target_list: 是目标列表，即含有重复数据的列表
    start_row: 是开始行，即工作表中开始比对数据的行（需要将标题除开）
    col: 是需要处理数据的列
    '''
    start = 0  # 开始行计数，初试值为0，对应列表中的第1个元素的位置0
    end = 0  # 结束行计数，初试值为0，对应列表中的第1个元素的位置0
    reference = target_list[0]  # 设定基准，以列表中的第一个字符串开始
    for i in range(len(target_list)):  # 遍历列表
        if target_list[i] != reference:  # 开始比对，如果内容不同执行如下
            reference = target_list[i]  # 基准变成列表中下一个字符串
            end = i - 1  # 列计数器
            ws.merge_cells(col + str(start + start_row) + ":" + col + str(end + start_row))
            start = end + 1
        if i == len(target_list) - 1:  # 遍历到最后一行，按如下操作
            end = i
            ws.merge_cells(col + str(start + start_row) + ":" + col + str(end + start_row))



def check_invoice (scode_database):
    data_hscode = pd.read_table(scode_database,sep='\t')
    source = filedialog.askopenfilename()
    souce_dossirt_name = source.split(".")[0].split("/")[-1]
    souce_path = str(source.split(".")[0]).replace(souce_dossirt_name,"")
    study_name = souce_path + souce_dossirt_name + " 海关码分析.xlsx"
    study_invoice(data_hscode,source,study_name)


def make_invoice(template,company_info):
    print("   ")
    print("***注意事项***")
    print("清关材料生成已清关数据VAT列为基准。一个VAT号码，一份清关材料。如果需要一个VAT生成两个甚至更多清关材料时"
          "VAT 后面可加入 - 1，-2 等让程序进行有效区分，"
          "生成清关材料之后，人工删掉相关多余信息")
    print("***注意事项***")
    print("")

    source = filedialog.askopenfilename()
    datasender = get_company_information(company_info)
    print(" Please find All sender information ")
    print(datasender[["发件人代码", "发件人英文"]])
    choose = input("Please choose the sender :", )
    choose_sender = datasender["发件人英文"].loc[datasender["发件人代码"] == int(choose)].tolist()[0]
    Nameofsender = choose_sender
    sender_adresse_complete = datasender["完整地址"].loc[datasender["发件人代码"] == int(choose)].tolist()[0]
    Sendercountrycode = datasender["国家代码"].loc[datasender["发件人代码"] == int(choose)].tolist()[0]
    Streetsender = datasender["地址"].loc[datasender["发件人代码"] == int(choose)].tolist()[0]
    Citysender = datasender["城市"].loc[datasender["发件人代码"] == int(choose)].tolist()[0]
    Senderzipcode = str(datasender["邮编"].loc[datasender["发件人代码"] == int(choose)].tolist()[0]).split(".")[0]
    print("")
    print("You have choosen :", choose_sender)
    print("Adresse complet is : ", sender_adresse_complete)
    print("")
    print("")
    datainvoice = get_invoicedate(source)
    vats = list(set(datainvoice["VAT号"].tolist()))
    vats.sort()
    ltas = list(set(datainvoice["提单号"].tolist()))
    if len(ltas) == 1:
        lta = ltas[0]
    else:
        lta = str(ltas)
    fuzhu = str(source).split("/")[-1]
    dir_name = source.replace(fuzhu,"") + lta + "--清关资料"
    if not os.path.isdir(dir_name):
        os.makedirs(dir_name)

    kg_brut_total = datainvoice['货箱重量(KG)'].sum()

    print("备注")
    print(" - 欧洲境内运费为必填（默认1300欧），国际运费为选填")

    transport_fee_interne = input("Please input the European shipping costs :", )
    transport_fee_externe = input("Please input the International Shipping Fees:", )

    try:
        transport_fee_interne = float(transport_fee_interne)
    except:
        transport_fee_interne = 1300

    print("")
    print("")
    print("info --- 正在生成提单    : " , lta)
    print("info --- 清关材料保存位置 : ",dir_name)
    print("info --- 共计清关材料数量 : ", len(vats), " Docs")
    print("info --- 包裹总数       : ",len(set(datainvoice["货箱编号"].tolist())), " Cartons")
    print("info --- 包裹总重       : ",kg_brut_total, " KG")

    if len(datainvoice)<1980:
        pass
    else:
        print("发票模板仅可容纳1980行数据，超过1980行，会被程序删掉。请联系相关人员进行处理")

    print("")
    print("")
    dic_lta = []
    a = 0
    for vat in vats:
        print("info --- 正在生成 vat：",vat)
        a = a + 1

        datainvoice_vat = datainvoice.loc[datainvoice['VAT号'] == vat]
        #获取交货条款;交货城市;清关方式;收件人国家
        incoterme = list(set(datainvoice_vat["交货条款"].tolist()))[0]
        incoterme_city = list(set(datainvoice_vat["交货城市"].tolist()))[0]
        delivery_country = list(set(datainvoice_vat["收件人国家"].tolist()))[0]
        code_regime = list(set(datainvoice_vat["清关方式"].tolist()))[0]

        qty_carton = len(set(datainvoice_vat["货箱编号"].tolist()))
        dir_name_hbl = dir_name+ "/" + lta + " --- " + vat + " - " + str(qty_carton)+"件 （ HBL " + str(a) + ")/"
        if not os.path.isdir(dir_name_hbl):
            os.makedirs(dir_name_hbl)
        #target = dir_name+ "/" + lta + " --- " + vat + " - " + str(qty_carton)+"件 （ HBL " + str(a) + ")/" + lta + " - INV&PL- " + vat + " - " + str(qty_carton)+"件 ( HBL " + str(a) + ").xlsx"
        target = dir_name_hbl + lta + " - INV&PL- " + vat + " - " + str(qty_carton) + "件 ( HBL " + str(a) + ").xlsx"
        copyfile(template, target)
        exporter_chi = "---"
        exporter_eng = Nameofsender
        ref_number = lta + " - " + str(a)
        invoice_number = "HBL - " + lta + " - " + str(a)
        importer = datainvoice_vat["收件人"].tolist()[0]
        EORI = datainvoice_vat["EORI"].tolist()[0]
        adresse = datainvoice_vat["地址"].tolist()[0]
        code_postal = str(datainvoice_vat["邮编"].tolist()[0]).split(".")[0]
        city = datainvoice_vat["城市"].tolist()[0]
        county_2_chiffre = datainvoice_vat["国家代码"].tolist()[0]
        county_complet = datainvoice_vat["国家全称"].tolist()[0]
        adresse_importer_complet = adresse + " ," + str(code_postal)+ " ," + city+ " ," + str(county_complet)
        #incoterme = str(incoterme)[:3] + " " + ville
        wb = load_workbook(target)
        invoice_sheet = wb.worksheets[0]
        # 填写excel invoice 表头信息
        invoice_sheet.cell(1, 1, exporter_chi)  # 出口商公司
        invoice_sheet.cell(2, 1, exporter_eng)  # 出口商英文
        invoice_sheet.cell(4, 3, exporter_eng)  # 出口商英文
        invoice_sheet.cell(5, 3, ref_number)  # 分单号
        invoice_sheet.cell(6, 3, Streetsender)  # 地址
        invoice_sheet.cell(7, 3, str(Senderzipcode))  # 邮编
        invoice_sheet.cell(8, 3, Citysender)  # 邮编
        invoice_sheet.cell(9, 3, Sendercountrycode)  # 邮编
        invoice_sheet.cell(4, 10, invoice_number)  # 发票号码
        invoice_sheet.cell(5, 10, date_now)  # 发日期

        # 填写进口商信息 excel invoice 表头信息
        invoice_sheet.cell(11, 3, importer)  # 进口商公司名称
        invoice_sheet.cell(12, 3, "")  # 电话
        invoice_sheet.cell(13, 3, adresse)  # 地址
        invoice_sheet.cell(14, 3, code_postal)  # 邮编
        invoice_sheet.cell(15, 3, city)  # 城市
        invoice_sheet.cell(16, 3, county_2_chiffre)  # 国家
        invoice_sheet.cell(17, 3, delivery_country)  # 收货国家
        invoice_sheet.cell(11, 10, vat)
        invoice_sheet.cell(12, 10, EORI)
        invoice_sheet.cell(13, 10, "EUR")  # 币种
        invoice_sheet.cell(14, 10, incoterme)
        invoice_sheet.cell(15, 10, incoterme_city)
        invoice_sheet.cell(16, 10, code_regime)  # 递延
        invoice_sheet.cell(17, 10, "")  # 邮箱
        wb.save(target)
        # 填写excel invoice 主体信息
        datainvoice_vat_traiter = datainvoice_vat[
            ["产品英文品名", "产品海关编码", '产品申报单价', '产品中文品名', '材质（须填写英文）', '货箱编号', '产品申报数量', '申报总价', '包裹净重', '货箱重量(KG)', "产品销售链接",
             "运单号"]]
        for x in range(len(datainvoice_vat_traiter)):
            for y in range(0,12):
                column = datainvoice_vat_traiter.columns[y]
                line = 20 + int(x)
                valeur = datainvoice_vat_traiter[column].tolist()[x]
                invoice_sheet.cell(line, y + 2, valeur)
                invoice_sheet.cell(line, y + 2).border = border
                invoice_sheet.cell(line, y + 2).alignment = align
            # 合并单元格
        #marks_list = []
        #tracking_list = []
        shipement_list = []
        for row in range(20, line + 1):
            #marks = invoice_sheet['G' + str(row)].value
            shipement = invoice_sheet['M' + str(row)].value
            #marks_list.append(marks)
            shipement_list.append(shipement)
        # 调用以上定义的合并单元格函数`Merge_cells`做单元格合并操作
        start_row = 20  # 开始行是第20行
        #Merge_cells(invoice_sheet, marks_list, start_row, "G")
        Merge_cells(invoice_sheet, shipement_list, start_row, "M")  # "M" - 票在最后一列
        # 填写excel invoice 结尾西信息
        invoice_sheet.delete_rows(line + 1, 2000 - line - 1)
        invoice_sheet.merge_cells(start_row=line + 1, start_column=3, end_row=line + 1, end_column=5)
        sum_pcs = datainvoice_vat_traiter['产品申报数量'].sum()
        invoice_sheet.cell(line + 1, 8, sum_pcs)
        sum_total_value = datainvoice_vat_traiter['申报总价'].sum()
        invoice_sheet.cell(line + 1, 9, sum_total_value)
        sum_total_net = datainvoice_vat_traiter['包裹净重'].sum()
        invoice_sheet.cell(line + 1, 10, sum_total_net)
        sum_total_brut = datainvoice_vat_traiter['货箱重量(KG)'].sum()
        invoice_sheet.cell(line + 1, 11, sum_total_brut)
        invoice_sheet.cell(line + 6, 3, round(transport_fee_interne * (sum_total_brut/kg_brut_total)))


        wb.save(target)
        # 处理 packing list
        pl_sheet = wb.worksheets[1]
        pl_sheet.cell(1, 1, exporter_chi)  # 发票抬头
        pl_sheet.cell(2, 1, exporter_eng)  # 发票英文名称
        pl_sheet.cell(4, 2, invoice_number)  # 发票英文名称
        data_pl_vat_traiter = datainvoice_vat[
            ["产品英文品名", '产品申报单价', '产品中文品名', '货箱编号', '产品净重', '箱数', '产品申报数量', '包裹净重', '货箱重量(KG)']]
        for x in range(len(data_pl_vat_traiter)):
            for y in range(9):
                column = data_pl_vat_traiter.columns[y]
                line = 6 + int(x)
                valeur = data_pl_vat_traiter[column].tolist()[x]
                pl_sheet.cell(line, y + 2, valeur)
                pl_sheet.cell(line, y + 2).border = border
                pl_sheet.cell(line, y + 2).alignment = align
        # 合单元格
        #marks_list = []  # 唛头
        carton_list = []  # 包裹数量
        for row in range(6, line+1):
            #marks = pl_sheet['E' + str(row)].value
            carton = pl_sheet['G' + str(row)].value
            #marks_list.append(marks)
            carton_list.append(carton)
        # 调用以上定义的合并单元格函数`Merge_cells`做单元格合并操作
        start_row = 6  # 开始行是第六行
        #Merge_cells(pl_sheet, marks_list, start_row, "E")  # "E" - 唛头是a列
        Merge_cells(pl_sheet, carton_list, start_row, "G")  # "G" - 箱数是在C列
        pl_sheet.delete_rows(line + 1, 2000 - line - 1)
        pl_sheet.cell(line + 1, 8, sum_pcs)
        pl_sheet.cell(line + 1, 7, qty_carton)
        pl_sheet.cell(line + 1, 9, sum_total_net)
        pl_sheet.cell(line + 1, 10, sum_total_brut)
        for row in range(6, line + 1):
            try:
                value = str(pl_sheet.cell(row, 7).value)
                if len(value) > 6:
                    pl_sheet.cell(row, 7, 1)
                else:
                    pl_sheet.cell(row, 7, 0)
            except:
                pass
        wb.save(target)

        # 处理 resume
        resume_sheet = wb.worksheets[2]
        datainvoice_vat_resume = datainvoice_vat_traiter.groupby(by="产品海关编码", sort=True).sum()
        nb_hscode = len(datainvoice_vat_resume)
        descriptions_hbl = list(set(datainvoice_vat_traiter["产品英文品名"].tolist()))
        descriptions_hbl.sort()
        descriptions_hbl = str(descriptions_hbl).replace("{", '').replace("}", '').replace("'", '').replace("[", '').replace("]", '')
        for x in range(len(datainvoice_vat_resume)):
            data = datainvoice_vat_resume[x:x + 1]
            hscode = data.index[0]
            descriptions = list(set(datainvoice_vat_traiter["产品英文品名"].loc[datainvoice_vat_traiter["产品海关编码"]==hscode].tolist()))
            descriptions.sort()
            descriptions = str(descriptions).replace("{", '').replace("}", '').replace("'", '').replace("[", '').replace("]", '')
            qty_hscode = data["产品申报数量"].tolist()[0]
            value_hscode = data["申报总价"].tolist()[0]
            kgnet_hscode = data["包裹净重"].tolist()[0]
            kgbrut_hscode = data["货箱重量(KG)"].tolist()[0]
            resume_sheet.cell(x + 2, 1, str(hscode))
            resume_sheet.cell(x + 2, 1).border = border
            resume_sheet.cell(x + 2, 2, qty_hscode)
            resume_sheet.cell(x + 2, 2).border = border
            resume_sheet.cell(x + 2, 3, value_hscode)
            resume_sheet.cell(x + 2, 3).border = border
            resume_sheet.cell(x + 2, 4, kgnet_hscode)
            resume_sheet.cell(x + 2, 4).border = border
            resume_sheet.cell(x + 2, 5, kgbrut_hscode)
            resume_sheet.cell(x + 2, 5).border = border
            resume_sheet.cell(x + 2, 6, descriptions)
            resume_sheet.cell(x + 2, 6).border = border

        resume_sheet.delete_rows(len(datainvoice_vat_resume) + 2, 1000 - len(datainvoice_vat_resume) - 2)
        resume_sheet.cell(x + 3, 2, datainvoice_vat_traiter["产品申报数量"].sum())
        resume_sheet.cell(x + 3, 3, datainvoice_vat_traiter["申报总价"].sum())
        resume_sheet.cell(x + 3, 4, datainvoice_vat_traiter["包裹净重"].sum())
        resume_sheet.cell(x + 3, 5, datainvoice_vat_traiter["货箱重量(KG)"].sum())
        wb.save(target)
        dic_resume = {"提单":lta,
                      "分单":lta+"-HBL-"+ str(a),
                      "税号":vat,
                      "包裹数量":qty_carton,
                      "净重":sum_total_net,
                      "毛重":sum_total_brut,
                      "海关码数量":nb_hscode,
                      "申报金额":sum_total_value,
                      "Description":descriptions_hbl,
                      "Company Trading":Nameofsender,
                      "Adresse shiper 1":Streetsender,
                      "Adresse shiper 2" :str(Citysender) + " " + str(Senderzipcode)+ " " + str(Sendercountrycode),
                      "Consignee":importer,
                      "Adresse Consignee 1":adresse,
                      "Adresse Consignee 2":city + " " + code_postal + " " + county_complet,
                      "CBM":"",
                      "Place of recepit":"",
                      "Port of loading":"",
                      "Ocean Vessel":"",
                      "Port of discharge":"",
                      "SealNo":"",
                      "Type":"",
                      "Rate":"",
                      "Prepaid at":"",
                      "Place Date Issu":""}
        dic_lta.append(dic_resume)
        # 开始生成 Begate文件
        dic_file = []
        list_hscode = list(set(datainvoice_vat["产品海关编码"].tolist()))
        list_hscode.sort()
        op = 1
        for hscode in list_hscode:
            description = set(datainvoice_vat["产品英文品名"].loc[datainvoice_vat["产品海关编码"] == hscode].tolist())
            Gooddescription = str(description).replace("{", '').replace("}", '').replace("'", '')
            Typeofpackages = "PC"
            Numberofpackages = datainvoice_vat["产品申报数量"].loc[datainvoice_vat["产品海关编码"] == hscode].sum()
            Brand_Marks = ""
            Netweight = datainvoice_vat["包裹净重"].loc[datainvoice_vat["产品海关编码"] == hscode].sum()
            Grossweight = datainvoice_vat["货箱重量(KG)"].loc[datainvoice_vat["产品海关编码"] == hscode].sum()
            Value = datainvoice_vat["申报总价"].loc[datainvoice_vat["产品海关编码"] == hscode].sum()
            Countryoforigin = "CN"
            Nameofsender = Nameofsender
            Streetsender = Streetsender
            Citysender = Citysender
            Senderzipcode = Senderzipcode
            Sendercountrycode = Sendercountrycode
            EORIsender = ""
            Nameofconsignee = importer
            Streetconsignee = adresse
            Cityofconsignee = city
            Zipcodeconsignee = code_postal
            Countrycodeconsignee = county_2_chiffre
            Track_Trace = op
            op = op + 1
            codeadditionnel = ""
            Invoicecurrency = "EUR"
            Incoterm = incoterme  # 这里注意
            countrycodeofdestination = delivery_country
            consigneeID = EORI
            if "GVR" in str(code_regime):
                code_regime = "4000"
            else:
                code_regime = code_regime
            or_4000_4200 = code_regime
            dic_hscode = {"HSCode": hscode,
                          "Gooddescription": Gooddescription,
                          "Typeofpackages": Typeofpackages,
                          "Numberofpackages": Numberofpackages,
                          "Brand_Marks": Brand_Marks,
                          "Netweight": Netweight,
                          "Grossweight": Grossweight,
                          "Value": Value,
                          "Countryoforigin": Countryoforigin,
                          "Nameofsender": Nameofsender,
                          "Streetsender": Streetsender,
                          "Citysender": Citysender,
                          "Senderzipcode": Senderzipcode,
                          "Sendercountrycode": Sendercountrycode,
                          "EORIsender": EORIsender,
                          "Nameofconsignee": Nameofconsignee,
                          "Streetconsignee": Streetconsignee,
                          "Cityofconsignee": Cityofconsignee,
                          "Zipcodeconsignee": Zipcodeconsignee,
                          "Countrycodeconsignee": Countrycodeconsignee,
                          "Track_Trace": Track_Trace,
                          "codeadditionnel": codeadditionnel,
                          "Invoicecurrency": Invoicecurrency,
                          "Incoterm": Incoterm,
                          "countrycodeofdestination":delivery_country ,
                          "consigneeID": consigneeID,
                          "or_4000_4200": or_4000_4200}
            dic_file.append(dic_hscode)
        df_begate = pd.DataFrame(list(dic_file))
        begate_name = dir_name_hbl + lta + " - BEGATE- " + vat + " (HBL " + str(a) + ").xlsx"
        df_begate.to_excel(begate_name, sheet_name='Begate file', index=False)
        print("info --- 税号 : ",vat,"，包裹数量 : ",qty_carton," --- 清关材料和Begate材料生成完毕")
        print("")
    dic_resume = {"提单":"共计",
                      "分单":"",
                      "税号":"",
                      "包裹数量":len(set(datainvoice["货箱编号"].tolist())),
                      "净重":datainvoice["包裹净重"].sum(),
                      "毛重":datainvoice["货箱重量(KG)"].sum(),
                      "海关码数量":"",
                      "申报金额":datainvoice["申报总价"].sum()}
    dic_lta.append(dic_resume)

    df_lta = pd.DataFrame(list(dic_lta))
    df_lta_name = dir_name + "/ " + lta + " 税号信息总结.xlsx"
    df_lta.to_excel(df_lta_name, sheet_name='税号信息总结', index=False)
    print("")
    print("")
    print("")
    print("info --- 税号信息文件储存位置：", df_lta_name)


print("Please choose your option:")

print(" 1. 检查核实清关材料")
print(" 2. 生成清关材料")

path_dossier = os.getcwd()

template = path_dossier +r"\template.xlsx"
company_info = path_dossier +"\invoice_tete.xlsx"
scode_database = path_dossier +"\hscode_database.txt"


option = input("Please input the number of your choice :", )
if str(option).replace(" ", "") == "1":
    check_invoice(scode_database)
elif str(option).replace(" ", "") == "2":
    make_invoice(template,company_info)
else:
    print("Sorry, please restart")
